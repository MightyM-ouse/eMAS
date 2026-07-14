#!/usr/bin/env python3
"""Generate the controlled eMAS report-redesign v3.2 templates and JSON contracts.

The generator contains only structural/report-binding metadata. It does not contain
business rules, RAG logic, customer data, confidential rates or sample records.
"""
from __future__ import annotations

import json
import shutil
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

ROOT = Path(__file__).resolve().parents[2]
TEMPLATE_VERSION = "1.2.0"
MAP_VERSION = "2.0.0"
RESULT_VERSION = "3.2.0"

NAVY = "17365D"
BLUE = "2F75B5"
LIGHT_BLUE = "D9EAF7"
PALE = "F3F6F9"
EDITABLE = "FFF2CC"
WHITE = "FFFFFF"
GREEN = "C6EFCE"
AMBER = "FFEB9C"
RED = "FFC7CE"
UNKNOWN = "E7E6E6"
THIN = Side(style="thin", color="B7C9D6")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CONTROLLED_LISTS = {
    "YesNo": ["Yes", "No"],
    "EvaluationStatus": ["Evaluated", "Warning", "Failed", "Not Assessed", "Not Applicable"],
    "RAG": ["Green", "Amber", "Red", "Unknown", "Not Assessed", "Information"],
    "Confidence": ["High", "Medium", "Low", "Unknown"],
    "ValueSource": ["Observed", "CustomerProvided", "Imported", "Derived", "Assumed", "NotAssessed"],
    "Priority": ["Critical", "High", "Medium", "Low", "Information"],
    "Severity": ["Critical", "High", "Medium", "Low", "Information"],
    "WorkflowStatus": ["Open", "In Progress", "Pending Review", "Closed", "Not Applicable"],
    "ReadinessResult": ["Ready", "Ready with Accepted Exceptions", "Blocked"],
    "VerificationResult": ["Reconciled", "Reconciled with Accepted Exceptions", "Not Reconciled", "Verification Incomplete"],
    "ReconciliationStatus": ["Matched", "Matched Within Tolerance", "Accepted Difference", "Review Required", "Mismatched", "Missing After Migration", "Unexpected After Migration", "Not Compared", "Evidence Missing"],
    "ReviewerDisposition": ["Confirmed Match", "Accepted Difference", "Remediation Required", "Further Investigation Required", "Not Applicable", "Duplicate Discrepancy", "Deferred with Approval", "Rejected", "Closed"],
}

SUMMARY_COLUMNS = ["Section", "Metric", "Value", "EvaluationStatus", "RAG", "Confidence", "ValueSource", "ReviewRequired", "Description"]
CONTROL_COLUMNS = ["ControlKey", "ControlValue", "Notes"]

PRE_SALES_SHEETS = [
    "01_Executive_Estimate", "02_Dossier_Inventory", "03_Sequence_Inventory", "04_Path_&_Volume_Inventory"
]
PRE_MIGRATION_SHEETS = [
    "01_Executive_Summary", "02_Readiness_Decision", "03_Inputs_Access_&_Transfer",
    "04_Dossier_Baseline", "05_Sequence_Baseline", "06_File_Type_Breakdown",
    "07_File_XML_Path_Checks", "08_Findings_&_Actions", "09_Exceptions_&_Exclusions",
    "10_Assumptions_&_Limits", "11_Execution_Details",
]
POST_MIGRATION_SHEETS = [
    "01_Executive_Summary", "02_Verification_Scope", "03_Overall_Reconciliation",
    "04_Dossier_Before_&_After", "05_Sequence_Before_&_After", "06_File_Type_&_Size_Comparison",
    "07_Database_Dossier_Inventory", "08_Import_Evidence_Review", "09_Discrepancies_&_Actions",
    "10_Accepted_Exceptions", "11_Assumptions_&_Limits", "12_Review_&_Execution",
    "Import Report Detail", "Post Import Verification", "Database Dossier Extract",
]

PS_DOSSIER = [
    "DossierId","Product","DossierDisplayName","DossierPath","SourceApplication","SourceApplicationVersion",
    "Region","RegionEvidence","RegionRuleIds","Authority","TechnicalStandard","Format","FormatEvidence","FormatRuleIds",
    "RegionalImplementation","PrimaryDossierType","ProductDomain","TypeEvidence","TypeRuleIds","ClassificationConfidence",
    "SequenceCount","SizeBytes","DisplaySizeGB","FileCount","FolderCount","EvaluationStatus","ValueSource","DossierRAG",
    "PrimaryRAGReason","ManualReviewRequired","ReviewReason","MigrationWorkstream","MigrationMethod","BatchEligible",
    "IndividualImportRequired","ConversionRequired","UpgradeDependency","EffortImpact","RecommendedAction","Comments",
]
PS_SEQUENCE = [
    "SequenceId","Product","DossierId","DossierDisplayName","SequenceDisplayName","SequencePath","Region","Authority",
    "TechnicalStandard","Format","RegionalImplementation","ClassificationConfidence","SizeBytes","DisplaySizeMB","FileCount",
    "FolderCount","BackboneXmlPresent","ChecksumIndicatorPresent","ZeroByteFileCount","LongestPathLength","EvaluationStatus",
    "ValueSource","SequenceRAG","PrimaryRAGReason","ReviewRequired","MigrationWorkstream","MigrationMethod","BatchEligible",
    "IndividualImportRequired","ConversionRequired","EffortImpact","RecommendedAction","Comments",
]
PS_EXPORT = [
    "EvidenceId","ExportRoot","Accessible","SizeBytes","DisplaySizeGB","DisplaySizeTB","FileCount","FolderCount",
    "DossierCount","SequenceCount","LongPathCount","ValueSource","EvaluationStatus","RAG","PrimaryRAGReason",
    "ReviewRequired","Comments",
]
PS_DIRECT = [
    "EvidenceId","SourceType","SourceReference","ResolvedPath","Accessible","SizeBytes","DisplaySizeGB","DisplaySizeTB",
    "ValueSource","IncludedInScope","EvaluationStatus","RAG","ReviewRequired","Comments",
]

PM_DECISION = [
    "DecisionId","ReadinessResult","DecisionConfidence","DecisionRationale","BlockerCount","WarningCount",
    "AcceptedExceptionCount","ExclusionCount","ReviewPendingCount","BlockerOverrideResult","RequiredNextAction","Owner",
    "TargetDate","EvaluationStatus","RAG","ValueSource","ReviewRequired","PolicyRuleIds","ReviewStatus","Comments",
]
PM_INPUT = [
    "CheckId","Area","RequirementType","EvidenceReference","Mandatory","Applicable","AccessibleAvailable","ObservedValue",
    "RequiredValue","Unit","ValueSource","EvaluationStatus","RAG","PrimaryRAGReason","Severity","Blocker","FindingCode",
    "ActionCode","ReviewRequired","Owner","TargetDate","Status","Comments",
]
PM_DOSSIER = [
    "DossierBaselineId","StableComparisonId","Product","DossierDisplayName","SourcePathReference","Region","Authority",
    "TechnicalStandard","RegionalImplementation","ProductDomain","LifecycleContext","ProductClass","ProcedureContext",
    "SourcePresentation","ClassificationEvidence","ClassificationRuleIds","ClassificationConfidence","SequenceCount","FileCount",
    "FolderCount","SizeBytes","DisplaySizeGB","FirstSequence","LastSequence","EvaluationStatus","DossierRAG","PrimaryRAGReason",
    "FindingCodes","RuleIds","Blocker","ReviewRequired","MigrationWorkstream","MigrationMethod","MigrationWave","ReadinessAction",
    "BaselineApprovalStatus","CarryForwardExceptionIds","ValueSource","Confidence","Comments",
]
PM_SEQUENCE = [
    "SequenceBaselineId","StableComparisonId","DossierBaselineId","Product","DossierDisplayName","SequenceDisplayName",
    "SequenceNumber","SequencePathReference","Region","Authority","TechnicalStandard","RegionalImplementation","FileCount",
    "FolderCount","SizeBytes","DisplaySizeMB","BackboneXmlStatus","ChecksumStatus","ReferencedFileStatus","EvaluationStatus",
    "SequenceRAG","PrimaryRAGReason","FindingCodes","RuleIds","Blocker","ReviewRequired","MigrationMethod","MigrationWave",
    "ReadinessAction","BaselineApprovalStatus","CarryForwardExceptionIds","ValueSource","Confidence","Comments",
]
PM_FILE_TYPE = [
    "RecordId","DossierId","DossierName","SequenceId","SequenceName","FileExtension","FileCategory","FileCount",
    "TotalSizeBytes","DisplaySizeMB","LargestFileSizeMB","ZeroByteCount","UnreadableFileCount","UnsupportedFileCount",
    "EncryptedFileCount","EvaluationStatus","RAG","PrimaryRAGReason","FindingCodes","RuleIds","ReviewRequired",
    "RecommendedAction","Comments",
]
PM_CHECKS = [
    "CheckId","EntityType","DossierId","SequenceId","CheckArea","CheckCode","RuleId","FindingCode","EvidenceReference",
    "ObservedValue","ExpectedValue","EvaluationStatus","RAG","Severity","Blocker","PrimaryRAGReason","ActionCode",
    "RecommendedAction","AcceptedException","ExceptionId","ReviewRequired","Owner","TargetDate","Status","Comments",
]
FINDINGS = [
    "FindingId","FindingCode","EntityType","EntityId","Area","EvidenceReference","ObservedValue","EvaluationStatus",
    "OriginalRAG","Severity","Blocker","PrimaryRAGReason","RuleIds","ReviewRequired","Status","Comments",
]
ACTIONS = [
    "ActionId","FindingId","ActionCode","ActionOrder","CustomerFacingAction","ConsultantFacingNote","Owner","TargetDate",
    "Status","ClosureEvidenceReference","ClosedBy","ClosedDate","Comments",
]
EXCEPTIONS = [
    "ExceptionId","FindingId","EntityType","EntityId","OriginalEvaluationStatus","OriginalRAG","OriginalEvidenceReference",
    "ExceptionPolicyId","PermittedEffect","ApprovalRole","ApprovalReference","ApprovalDate","SupportingEvidenceReference",
    "ValidFrom","ValidUntil","CarryForwardRequired","ReviewStatus","Comments",
]
EXCLUSIONS = [
    "ExclusionId","EntityType","EntityId","ScopeDescription","Reason","Authority","ApprovedBy","ApprovalDate","BaselineEffect",
    "PostMigrationTreatment","Comments",
]
ASSUMPTIONS = ["AssumptionId","Scope","Statement","Impact","ValueSource","Confidence","ReviewRequired","Owner","Status","Comments"]
LIMITATIONS = ["LimitationId","Scope","Statement","Impact","EvidenceReference","ValueSource","Confidence","ReviewRequired","Owner","Status","Comments"]

POST_SCOPE = [
    "ScopeItemId","ScopeArea","EvidenceType","EvidenceVersion","EvidenceGeneratedAt","EvidenceChecksum","BaselineSource",
    "ImportEvidenceSource","DatabaseEvidenceSource","PostImportEvidenceSource","ComparisonKey","MandatoryEvidence",
    "EvidenceAvailable","CompatibilityStatus","Included","ExclusionId","ExceptionId","EvidenceLimitation","ValueSource",
    "EvaluationStatus","RAG","Confidence","ReviewRequired","Comments",
]
POST_OVERALL = [
    "ReconciliationMetricId","Area","MetricCode","MetricName","EntityScope","BaselineValue","ImportReportValue","DatabaseValue",
    "PostImportValue","AbsoluteDifference","PercentageDifference","Unit","Tolerance","ComparisonResult","ReconciliationStatus",
    "EvaluationStatus","RAG","PrimaryRAGReason","AcceptedException","ExceptionId","ReviewRequired","DiscrepancyId",
    "SystemComparisonNote","ReviewerNote","ReviewerDisposition",
]
POST_DOSSIER = [
    "DossierComparisonId","StableComparisonId","BaselineDossierId","TargetDatabaseDossierId","Product","BaselineDossierName",
    "ImportedDossierName","DatabaseDossierName","PostImportDossierName","Region","Authority","TechnicalStandard",
    "RegionalImplementation","PrimaryDossierType","MigrationMethod","MigrationWave","BaselinePresent","ImportReportPresent",
    "DatabasePresent","PostImportEvidencePresent","BaselineSequenceCount","ImportReportSequenceCount","DatabaseSequenceCount",
    "PostImportSequenceCount","SequenceCountDifference","BaselineFileCount","PostImportFileCount","FileCountDifference",
    "BaselineFolderCount","PostImportFolderCount","FolderCountDifference","BaselineSizeBytes","PostImportSizeBytes","SizeDifferenceBytes",
    "ImportStatusSummary","PreMigrationRAG","PostMigrationRAG","PrimaryRAGReason","ReconciliationStatus","DiscrepancyIds",
    "AcceptedException","ExceptionIds","ReviewRequired","SystemComparisonNote","ReviewerNote","ReviewerDisposition","Owner",
    "TargetDate","ReviewStatus","RecommendedAction","Comments",
]
POST_SEQUENCE = [
    "SequenceComparisonId","StableComparisonId","BaselineDossierId","TargetDatabaseDossierId","DossierName","BaselineSequenceId",
    "TargetSequenceId","SequenceDisplayName","Region","TechnicalStandard","RegionalImplementation","MigrationMethod","MigrationWave",
    "BaselinePresent","ImportReportPresent","DatabasePresent","PostImportEvidencePresent","BaselineFileCount","PostImportFileCount",
    "FileCountDifference","BaselineFolderCount","PostImportFolderCount","FolderCountDifference","BaselineSizeBytes","PostImportSizeBytes",
    "SizeDifferenceBytes","BaselineBackboneXmlStatus","PostImportBackboneXmlStatus","BaselineChecksumStatus","PostImportChecksumStatus",
    "BaselineReferencedFileStatus","PostImportReferencedFileStatus","ImportStatus","ImportStatusCode","PreMigrationRAG","PostMigrationRAG",
    "PrimaryRAGReason","ReconciliationStatus","DiscrepancyIds","AcceptedException","ExceptionIds","ReviewRequired",
    "SystemComparisonNote","ReviewerNote","ReviewerDisposition","Owner","TargetDate","ReviewStatus","RecommendedAction","Comments",
]
POST_FILE_TYPE = [
    "ComparisonId","DossierId","DossierName","SequenceId","SequenceName","FileExtension","FileCategory","BaselineFileCount",
    "PostMigrationFileCount","FileCountDifference","BaselineTotalSizeBytes","PostMigrationTotalSizeBytes","SizeDifferenceBytes",
    "BaselineLargestFileSizeBytes","PostMigrationLargestFileSizeBytes","BaselineZeroByteCount","PostMigrationZeroByteCount",
    "BaselineUnsupportedCount","PostMigrationUnsupportedCount","BaselineEncryptedCount","PostMigrationEncryptedCount","Tolerance",
    "ReconciliationStatus","EvaluationStatus","RAG","PrimaryRAGReason","AcceptedException","ExceptionId","DiscrepancyId",
    "ReviewRequired","SystemComparisonNote","ReviewerNote","ReviewerDisposition","RecommendedAction","Comments",
]
POST_DATABASE = [
    "DatabaseRecordId","TargetDossierId","StableComparisonId","DossierName","Product","Region","AuthorityProcedure",
    "TechnicalStandard","RegionalImplementation","LifecycleStatus","SequenceCount","FirstSequence","LastSequence","CreationDate",
    "ImportDate","LastModifiedDate","ActiveDeletedStatus","ArchiveRepositoryReference","SourceQueryExtractReference","ValueSource",
    "EvaluationStatus","Confidence","BaselineMatchStatus","RAG","PrimaryRAGReason","ReviewRequired","ReviewerNote","Comments",
]
POST_IMPORT = [
    "ImportEvidenceId","SourceRowReference","SourceName","SourceDossier","SourceSequence","DestinationDossier","DestinationSequence",
    "ImportStartTime","ImportEndTime","Duration","RawStatus","RawStatusCode","RawMessage","NormalizedCategory","NormalizedResult",
    "EvaluationStatus","RAG","PrimaryRAGReason","RuleId","FindingCode","DiscrepancyId","ReviewRequired","RecommendedAction",
    "ReviewerNote","ReviewerDisposition","Comments",
]
POST_DISCREPANCY = [
    "DiscrepancyId","EntityType","DossierId","DossierName","SequenceId","ComparisonArea","ExpectedValue","ImportValue","DatabaseValue",
    "ObservedValue","Difference","Classification","Severity","EvaluationStatus","RAG","PrimaryRAGReason","Blocker","AcceptedException",
    "ExceptionId","RequiredAction","Owner","TargetDate","SystemNote","ReviewerNote","ReviewerDisposition","ReviewStatus",
    "ClosureEvidenceReference","ClosedBy","ClosedDate","Comments",
]
POST_ACCEPTED = [
    "ExceptionId","OriginalFindingId","EntityType","EntityId","OriginalEvaluationStatus","OriginalRAG","OriginalEvidenceReference",
    "ExceptionPolicyId","ApprovedEffect","ApprovalReference","ApprovalDate","SupportingEvidenceReference","ValidUntil",
    "CarryForwardStatus","ObservedDifference","CoverageStatus","EvaluationStatus","CurrentRAG","ReviewerDisposition","Comments",
]
POST_RAW_IMPORT = ["Source.Name","Source.Dossier","Source.Sequence","Destination.Dossier","Destination.Sequence","Status","Status.Code","Message","Start.Time","End.Time"]
POST_RAW_VERIFY = ["DossierDirecotry","DossierName","Sequence","FileCount","FolderCount","SizeBytes","Region","Status","Comments"]
POST_RAW_DATABASE = [
    "DossierId","DossierName","Product","Region","Authority","TechnicalStandard","RegionalImplementation","LifecycleStatus",
    "SequenceCount","FirstSequence","LastSequence","CreationDate","ImportDate","LastModifiedDate","ActiveDeletedStatus",
    "ArchiveRepositoryReference",
]

PS_METRICS = [
    ("Assessment Context","Assessment Mode","Customer-selected collection mode."),
    ("Assessment Context","Current Application","Current source application or system."),
    ("Assessment Context","Current Version","Current version when known."),
    ("Assessment Context","Target Planning Status","Pending EXTEDO Review during customer collection."),
    ("Migration Approach","Recommended Migration Scenario","Completed during EXTEDO review."),
    ("Migration Approach","Required Workstreams","Ordered configured workstreams."),
    ("Scope and Workload","Total Dossiers","Export population only."),
    ("Scope and Workload","Total Sequences","Export population only."),
    ("Scope and Workload","Export Size","Observed export volume."),
    ("Scope and Workload","Archive Size","Aggregate direct-copy volume."),
    ("Scope and Workload","Index Size","Aggregate direct-copy volume."),
    ("Scope and Workload","Database Size","Aggregate direct-copy volume."),
    ("Scope and Workload","Green Dossiers","Evidence-quality population."),
    ("Scope and Workload","Amber Dossiers","Evidence-quality population."),
    ("Scope and Workload","Red Dossiers","Evidence-quality population."),
    ("Scope and Workload","Unknown Dossiers","Evidence-quality population."),
    ("Effort","Overall Complexity","Completed during EXTEDO review."),
    ("Effort","Estimated Technical Effort","Internal profile only."),
    ("Effort","Most Likely Estimate","Internal profile only."),
    ("Effort","Estimate Confidence","Confidence in estimate."),
    ("Effort","Primary Effort Drivers","Configured ordered drivers."),
    ("Next Steps","Required Decisions and Clarifications","Commercially or technically meaningful open items."),
]
PM_METRICS = [
    ("Assessment Context","Baseline ID","Attributable comparison baseline."),
    ("Readiness Decision","Readiness Result","Ready, Ready with Accepted Exceptions or Blocked."),
    ("Readiness Decision","Decision Confidence","Reliability of the decision."),
    ("Readiness Decision","Blocker Count","Unresolved blockers."),
    ("Readiness Decision","Warning Count","Recoverable warnings."),
    ("Readiness Decision","Accepted Exception Count","Approved residual exceptions."),
    ("Dossier Readiness","Baseline Dossiers","Approved dossier population."),
    ("Dossier Readiness","Green Dossiers","Ready evidence population."),
    ("Dossier Readiness","Amber Dossiers","Action or exception population."),
    ("Dossier Readiness","Red Dossiers","Blocked population."),
    ("Sequence Readiness","Baseline Sequences","Approved sequence population."),
    ("Sequence Readiness","Green Sequences","Ready evidence population."),
    ("Sequence Readiness","Amber Sequences","Action or exception population."),
    ("Sequence Readiness","Red Sequences","Blocked population."),
    ("Scope and Volume","Total Files","Approved baseline count."),
    ("Scope and Volume","Total Size","Approved baseline volume."),
    ("Migration Approach","Migration Methods","Approved method populations."),
    ("Migration Approach","Migration Waves","Approved wave populations."),
    ("Handover","Baseline Integrity","Integrity and compatibility status."),
    ("Handover","Post-Migration Carry-Forward","Baseline handover readiness."),
]
POST_METRICS = [
    ("Verification Context","Baseline Compatibility","Approved baseline compatibility."),
    ("Final Result","Verification Result","Reconciled, Reconciled with Accepted Exceptions, Not Reconciled or Verification Incomplete."),
    ("Final Result","Verification Confidence","Reliability of the conclusion."),
    ("Final Result","Unresolved Discrepancy Count","Open unresolved differences."),
    ("Dossier Reconciliation","Baseline Dossiers","Expected dossier population."),
    ("Dossier Reconciliation","Import Report Dossiers","Observed import-report population."),
    ("Dossier Reconciliation","Database Dossiers","Observed target-database population."),
    ("Dossier Reconciliation","Post-Import Dossiers","Observed post-import population."),
    ("Dossier Reconciliation","Missing Dossiers","Expected dossiers not observed."),
    ("Dossier Reconciliation","Unexpected Dossiers","Observed dossiers outside baseline."),
    ("Sequence Reconciliation","Baseline Sequences","Expected sequence population."),
    ("Sequence Reconciliation","Missing Sequences","Expected sequences not observed."),
    ("Sequence Reconciliation","Unexpected Sequences","Observed sequences outside baseline."),
    ("File and Volume","File Count Difference","Baseline versus post-import difference."),
    ("File and Volume","Size Difference","Baseline versus post-import difference."),
    ("Import Execution","Failed Imports","Normalized failed import rows."),
    ("Import Execution","Warning Imports","Normalized warning import rows."),
    ("Database Reconciliation","Missing Database Dossiers","Expected dossiers absent from target extract."),
    ("Review and Handover","Open Actions","Actions remaining before closeout."),
    ("Review and Handover","Closeout Evidence Ready","Technical closeout evidence status."),
]


def table(sheet, name, source, columns, header=5, write="appendRows", protected=None, metrics=None, start_col=1):
    return {
        "sheetName": sheet,
        "tableName": name,
        "sourceCollection": source,
        "writeMode": write,
        "headerRow": header,
        "firstDataRow": header + 1,
        "startColumn": start_col,
        "rowCapacity": {"maxPreProvisionedRows": 1000, "rowInsertionBehavior": "appendWithinPreProvisionedRange", "followingTable": None},
        "columns": [column_contract(c) for c in columns],
        **({"protection": protected} if protected else {}),
        **({"metricRows": metrics} if metrics else {}),
    }


def column_contract(name):
    lowered = name.lower()
    list_name = None
    if name in ("RAG", "DossierRAG", "SequenceRAG", "OriginalRAG", "CurrentRAG", "PostMigrationRAG", "PreMigrationRAG"):
        list_name = "RAG"
    elif name == "EvaluationStatus" or name == "OriginalEvaluationStatus":
        list_name = "EvaluationStatus"
    elif "Confidence" in name:
        list_name = "Confidence"
    elif name in ("ValueSource",):
        list_name = "ValueSource"
    elif name in ("ReviewRequired", "Mandatory", "Applicable", "Included", "Blocker", "AcceptedException", "CarryForwardRequired", "BatchEligible", "IndividualImportRequired", "ConversionRequired"):
        list_name = "YesNo"
    elif name in ("Severity",):
        list_name = "Severity"
    elif name in ("Priority",):
        list_name = "Priority"
    elif name in ("ReadinessResult",):
        list_name = "ReadinessResult"
    elif name in ("VerificationResult",):
        list_name = "VerificationResult"
    elif name in ("ReconciliationStatus",):
        list_name = "ReconciliationStatus"
    elif name in ("ReviewerDisposition",):
        list_name = "ReviewerDisposition"
    dtype = "String"
    if "date" in lowered or "time" in lowered:
        dtype = "DateTime"
    elif "count" in lowered or "bytes" in lowered or "size" in lowered or "difference" in lowered or "percentage" in lowered:
        dtype = "Number"
    elif name.endswith("Id") or name.endswith("Ids") or name.endswith("Code") or name.endswith("Codes"):
        dtype = "Identifier"
    elif any(token in lowered for token in ("reason", "rationale", "comment", "note", "action", "statement", "description", "message")):
        dtype = "LongText"
    return {"targetColumn": name, "sourceField": name[0].lower() + name[1:], "required": False, "dataType": dtype, "controlledValueList": list_name, "format": None}


def phase_specs():
    ps_tables = [
        table("01_Executive_Estimate", "tblPreSalesExecutiveEstimate", "executiveMetrics", SUMMARY_COLUMNS, metrics=PS_METRICS),
        table("01_Executive_Estimate", "tblPreSalesDecisionsClarifications", "decisionClarifications", ["Priority","Area","DecisionOrClarificationRequired","Reason","EffectOnEstimate","RequiredFrom","RequiredBeforeQuotation","Owner","TargetDate","Status","Note"], header=32),
        table("02_Dossier_Inventory", "tblPreSalesDossierInventory", "dossierInventory", PS_DOSSIER),
        table("03_Sequence_Inventory", "tblPreSalesSequenceInventory", "sequenceInventory", PS_SEQUENCE),
        table("04_Path_&_Volume_Inventory", "tblPreSalesExportEvidence", "exportEvidence", PS_EXPORT),
        table("04_Path_&_Volume_Inventory", "tblPreSalesDirectCopyEvidence", "directCopyEvidence", PS_DIRECT, start_col=20),
        table("04_Path_&_Volume_Inventory", "tblPreSalesTemplateControl", "templateControl", CONTROL_COLUMNS, header=20, write="staticReleaseManaged", protected="readOnly"),
    ]
    pm_tables = [
        table("01_Executive_Summary", "tblPreMigrationExecutiveSummary", "summaryMetrics", SUMMARY_COLUMNS, metrics=PM_METRICS),
        table("01_Executive_Summary", "tblPreMigrationCriticalActions", "criticalActions", ["Priority","FindingId","EntityType","EntityId","PrimaryRAGReason","RequiredAction","Owner","TargetDate","Status","ExceptionId"], header=30),
        table("02_Readiness_Decision", "tblPreMigrationReadinessDecision", "readinessDecision", PM_DECISION),
        table("03_Inputs_Access_&_Transfer", "tblPreMigrationInputChecks", "inputChecks", PM_INPUT),
        table("04_Dossier_Baseline", "tblPreMigrationDossierBaseline", "dossierBaseline", PM_DOSSIER),
        table("05_Sequence_Baseline", "tblPreMigrationSequenceBaseline", "sequenceBaseline", PM_SEQUENCE),
        table("06_File_Type_Breakdown", "tblPreMigrationFileTypeBreakdown", "fileTypeBreakdown", PM_FILE_TYPE),
        table("07_File_XML_Path_Checks", "tblPreMigrationFileXmlPathChecks", "fileXmlPathChecks", PM_CHECKS),
        table("08_Findings_&_Actions", "tblPreMigrationFindings", "findings", FINDINGS),
        table("08_Findings_&_Actions", "tblPreMigrationActions", "actions", ACTIONS, start_col=19),
        table("09_Exceptions_&_Exclusions", "tblPreMigrationAcceptedExceptions", "acceptedExceptions", EXCEPTIONS),
        table("09_Exceptions_&_Exclusions", "tblPreMigrationExclusions", "exclusions", EXCLUSIONS, start_col=22),
        table("10_Assumptions_&_Limits", "tblPreMigrationAssumptions", "assumptions", ASSUMPTIONS),
        table("10_Assumptions_&_Limits", "tblPreMigrationLimitations", "limitations", LIMITATIONS, start_col=13),
        table("11_Execution_Details", "tblPreMigrationExecutionDetails", "executionDetails", SUMMARY_COLUMNS),
        table("11_Execution_Details", "tblPreMigrationTemplateControl", "templateControl", CONTROL_COLUMNS, header=20, write="staticReleaseManaged", protected="readOnly"),
    ]
    post_tables = [
        table("01_Executive_Summary", "tblPostMigrationExecutiveSummary", "summaryMetrics", SUMMARY_COLUMNS, metrics=POST_METRICS),
        table("01_Executive_Summary", "tblPostMigrationCriticalDiscrepancies", "criticalDiscrepancies", ["Priority","DiscrepancyId","Area","DossierName","SequenceId","Difference","RAG","PrimaryRAGReason","AcceptedException","RequiredAction","Owner","TargetDate","ReviewStatus","ReviewerNote"], header=30),
        table("02_Verification_Scope", "tblPostMigrationVerificationScope", "verificationScope", POST_SCOPE),
        table("03_Overall_Reconciliation", "tblPostMigrationOverallReconciliation", "overallReconciliation", POST_OVERALL),
        table("04_Dossier_Before_&_After", "tblPostMigrationDossierBeforeAfter", "dossierBeforeAfter", POST_DOSSIER),
        table("05_Sequence_Before_&_After", "tblPostMigrationSequenceBeforeAfter", "sequenceBeforeAfter", POST_SEQUENCE),
        table("06_File_Type_&_Size_Comparison", "tblPostMigrationFileTypeSizeComparison", "fileTypeSizeComparison", POST_FILE_TYPE),
        table("07_Database_Dossier_Inventory", "tblPostMigrationDatabaseDossierInventory", "databaseDossierInventory", POST_DATABASE),
        table("08_Import_Evidence_Review", "tblPostMigrationImportEvidenceReview", "importEvidenceReview", POST_IMPORT),
        table("09_Discrepancies_&_Actions", "tblPostMigrationDiscrepanciesActions", "discrepanciesAndActions", POST_DISCREPANCY),
        table("10_Accepted_Exceptions", "tblPostMigrationAcceptedExceptions", "acceptedExceptions", POST_ACCEPTED),
        table("11_Assumptions_&_Limits", "tblPostMigrationAssumptionsLimitations", "assumptionsAndLimitations", ASSUMPTIONS + ["LimitationType","EvidenceReference"]),
        table("12_Review_&_Execution", "tblPostMigrationReviewExecution", "reviewAndExecution", SUMMARY_COLUMNS),
        table("12_Review_&_Execution", "tblPostMigrationTemplateControl", "templateControl", CONTROL_COLUMNS, header=20, write="staticReleaseManaged", protected="readOnly"),
        table("Import Report Detail", "tblRawImportReportDetail", "rawImportReportDetail", POST_RAW_IMPORT, protected="appendOnly"),
        table("Post Import Verification", "tblRawPostImportVerification", "rawPostImportVerification", POST_RAW_VERIFY, protected="appendOnly"),
        table("Database Dossier Extract", "tblRawDatabaseDossierExtract", "rawDatabaseDossierExtract", POST_RAW_DATABASE, protected="appendOnly"),
    ]
    return {
        "pre-sales": ("PRE_SALES", "Pre-Sales Assessment", "EMAS-TPL-PRESALES", "eMAS_PreSales_Template.xlsx", PRE_SALES_SHEETS, ps_tables),
        "pre-migration": ("PRE_MIGRATION", "Pre-Migration Readiness", "EMAS-TPL-PREMIGRATION", "eMAS_PreMigration_Template.xlsx", PRE_MIGRATION_SHEETS, pm_tables),
        "post-migration": ("POST_MIGRATION", "Post-Migration Verification", "EMAS-TPL-POSTMIGRATION", "eMAS_PostMigration_Template.xlsx", POST_MIGRATION_SHEETS, post_tables),
    }


def build_map(phase_key, spec):
    phase_code, phase_name, template_id, template_name, sheets, tables = spec
    protected = []
    for t in tables:
        if t.get("protection"):
            protected.append({
                "sheetName": t["sheetName"], "tableName": t["tableName"], "protection": t["protection"],
                "preserveHeadersExactly": t["protection"] == "appendOnly",
                "reason": "Release-managed control metadata." if t["protection"] == "readOnly" else "Raw evidence rows and literal headers are append-only and preserved exactly.",
            })
    path_segment = {"pre-sales":"pre-sales","pre-migration":"pre-migration","post-migration":"post-migration"}[phase_key]
    return {
        "mappingId": f"EMAS-MAP-{phase_code.replace('_','')}",
        "mappingVersion": MAP_VERSION,
        "phaseCode": phase_code,
        "phaseName": phase_name,
        "description": "Technical result-to-workbook binding for the approved report redesign v3.2. Contains no business rules, RAG logic, readiness logic, reconciliation logic or effort logic.",
        "template": {
            "templateId": template_id, "templateName": template_name, "templateVersion": TEMPLATE_VERSION,
            "templateFilePath": f"templates/controlled/{path_segment}/{template_name}",
            "templateControlTable": next(t["tableName"] for t in tables if t.get("protection") == "readOnly"),
        },
        "resultSchemaPath": f"config/result-schemas/report-redesign-v3.2/{phase_key}.result.schema.json",
        "requiredSheetOrder": sheets,
        "protectedTables": protected,
        "tableMappings": [{k:v for k,v in t.items() if k not in ("protection","metricRows")} for t in tables],
    }


def map_schema():
    return {
        "$schema":"https://json-schema.org/draft/2020-12/schema",
        "$id":"https://github.com/MightyM-ouse/eMAS/config/report-mappings/report-template-map.schema.json",
        "title":"eMAS Report Template Map v3.2",
        "type":"object","additionalProperties":False,
        "required":["mappingId","mappingVersion","phaseCode","phaseName","template","requiredSheetOrder","protectedTables","tableMappings"],
        "properties":{
            "mappingId":{"type":"string"},"mappingVersion":{"type":"string"},"phaseCode":{"enum":["PRE_SALES","PRE_MIGRATION","POST_MIGRATION"]},
            "phaseName":{"type":"string"},"description":{"type":"string"},"resultSchemaPath":{"type":"string"},
            "template":{"type":"object","required":["templateId","templateName","templateVersion","templateFilePath","templateControlTable"],"additionalProperties":True},
            "requiredSheetOrder":{"type":"array","items":{"type":"string"},"minItems":1,"uniqueItems":True},
            "protectedTables":{"type":"array","items":{"type":"object","additionalProperties":True}},
            "tableMappings":{"type":"array","minItems":1,"items":{"type":"object","required":["sheetName","tableName","sourceCollection","writeMode","headerRow","firstDataRow","columns"],"additionalProperties":True}},
        },
    }


def result_schema(phase_key, mapping):
    properties = {"phaseCode":{"const":mapping["phaseCode"]},"resultContractVersion":{"type":"string"}}
    required = ["phaseCode","resultContractVersion"]
    for source in dict.fromkeys(t["sourceCollection"] for t in mapping["tableMappings"] if t["sourceCollection"] != "templateControl"):
        properties[source] = {"type":"array","items":{"type":"object","additionalProperties":True}}
        required.append(source)
    if phase_key == "pre-sales":
        properties.update({
            "executionProfile":{"enum":["CustomerCollection","CustomerAssessment","InternalEstimation"]},
            "assessmentMode":{"type":"object","additionalProperties":True},
            "assessmentContext":{"type":"object","additionalProperties":True},
            "currentSystem":{"type":"object","additionalProperties":True},
            "targetPlanning":{"type":"object","additionalProperties":True},
            "migrationScenarioDecision":{"type":"object","additionalProperties":True},
            "effortEstimate":{"type":"object","additionalProperties":True},
        })
        required += ["executionProfile","assessmentMode","assessmentContext","currentSystem","targetPlanning","effortEstimate"]
        properties["directCopyEvidence"]["items"]["not"] = {"anyOf":[{"required":["fileCount"]},{"required":["folderCount"]},{"required":["longPathCount"]},{"required":["largestFile"]}]}
    return {"$schema":"https://json-schema.org/draft/2020-12/schema","title":f"eMAS {mapping['phaseName']} Result v3.2","type":"object","additionalProperties":False,"required":required,"properties":properties}


def display_width(header):
    h = header.lower()
    if any(x in h for x in ("description","reason","rationale","comment","note","action","statement","message")): return 34
    if any(x in h for x in ("path","reference","evidence","rule","finding","exception","discrepancy")): return 25
    if any(x in h for x in ("application","implementation","standard","method","status","date","time")): return 20
    if any(x in h for x in ("count","size","value","unit","rag","confidence","priority","severity")): return 16
    return min(max(len(header)+2, 14), 22)


def heading(ws, phase_name, sheet_name, raw=False):
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:H1"); ws["A1"] = f"eMAS — {phase_name}"
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY); ws["A1"].font = Font(color=WHITE,bold=True,size=16); ws.row_dimensions[1].height=28
    ws.merge_cells("A2:H2"); ws["A2"] = sheet_name.replace("_"," ")
    ws["A2"].fill = PatternFill("solid", fgColor=LIGHT_BLUE); ws["A2"].font = Font(color=NAVY,bold=True,size=12); ws.row_dimensions[2].height=22
    ws.merge_cells("A3:H3")
    ws["A3"] = ("Raw evidence: preserve source headers and rows exactly; interpretation belongs in the normalized sheets." if raw else "Controlled blank template aligned to Enterprise Requirements v3.2. Source evidence remains read-only and business interpretation is configuration-driven.")
    ws["A3"].fill = PatternFill("solid", fgColor=PALE); ws["A3"].font = Font(color="404040",italic=True,size=9); ws["A3"].alignment=Alignment(wrap_text=True,vertical="top"); ws.row_dimensions[3].height=30
    ws.freeze_panes = "A5"; ws.print_title_rows="$1:$5"; ws.page_setup.orientation="landscape"; ws.page_setup.fitToWidth=1; ws.sheet_properties.pageSetUpPr.fitToPage=True


def metric_rows(metrics, columns):
    rows=[]
    for section, metric, description in metrics:
        values={"Section":section,"Area":section,"Metric":metric,"EvaluationStatus":"Not Assessed","ValueSource":"Derived","ReviewRequired":"No","Description":description}
        rows.append([values.get(c,"") for c in columns])
    return rows


def control_rows(mapping, columns):
    items=[("TemplateId",mapping["template"]["templateId"]),("TemplateVersion",mapping["template"]["templateVersion"]),("PhaseCode",mapping["phaseCode"]),("Status","Controlled Working"),("RequiredSheetNames","; ".join(mapping["requiredSheetOrder"])),("RequiredTableNames","; ".join(t["tableName"] for t in mapping["tableMappings"]))]
    return [[k if c=="ControlKey" else v if c=="ControlValue" else "" for c in columns] for k,v in items]


def add_table(ws, table_def, mapping, metrics_lookup):
    columns=[c["targetColumn"] for c in table_def["columns"]]
    start_col=int(table_def.get("startColumn",1)); header=int(table_def["headerRow"]); first=int(table_def["firstDataRow"])
    metrics=metrics_lookup.get(table_def["tableName"])
    if metrics:
        rows=metric_rows(metrics,columns)
    elif table_def["sourceCollection"]=="templateControl":
        rows=control_rows(mapping,columns)
    else:
        rows=[["" for _ in columns]]
    for i,col in enumerate(columns,start_col):
        cell=ws.cell(header,i,col); cell.fill=PatternFill("solid",fgColor=BLUE); cell.font=Font(color=WHITE,bold=True,size=9); cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); cell.border=BORDER
        ws.column_dimensions[get_column_letter(i)].width=max(ws.column_dimensions[get_column_letter(i)].width or 0,display_width(col))
    ws.row_dimensions[header].height=32
    for r,row in enumerate(rows,first):
        for j,value in enumerate(row,start_col):
            cell=ws.cell(r,j,value); cell.border=BORDER; cell.alignment=Alignment(vertical="top",wrap_text=True)
            if value=="" and table_def["writeMode"]!="staticReleaseManaged": cell.fill=PatternFill("solid",fgColor=EDITABLE)
        ws.row_dimensions[r].height=22
    end=first+len(rows)-1
    ref=f"{get_column_letter(start_col)}{header}:{get_column_letter(start_col+len(columns)-1)}{end}"
    tbl=Table(displayName=table_def["tableName"],ref=ref); tbl.tableStyleInfo=TableStyleInfo(name="TableStyleMedium2",showRowStripes=True,showFirstColumn=False,showLastColumn=False,showColumnStripes=False); ws.add_table(tbl)
    for idx,col_def in enumerate(table_def["columns"],start_col):
        list_name=col_def.get("controlledValueList"); values=CONTROLLED_LISTS.get(list_name)
        if values:
            formula='"'+','.join(values)+'"'
            if len(formula)<=255:
                dv=DataValidation(type="list",formula1=formula,allow_blank=True); ws.add_data_validation(dv); letter=get_column_letter(idx); dv.add(f"{letter}{first}:{letter}{max(first+999,1000)}")
    if "RAG" in columns:
        letter=get_column_letter(start_col+columns.index("RAG")); end_row=max(first+999,1000)
        for value,color in (("Green",GREEN),("Amber",AMBER),("Red",RED),("Unknown",UNKNOWN),("Not Assessed",UNKNOWN)):
            ws.conditional_formatting.add(f"{letter}{first}:{letter}{end_row}",FormulaRule(formula=[f'{letter}{first}="{value}"'],fill=PatternFill("solid",fgColor=color)))


def generate_workbook(phase_key, mapping, metrics_lookup, output):
    wb=Workbook(); wb.remove(wb.active); wb.properties.creator="EXTEDO eMAS"; wb.properties.title=f"eMAS {mapping['phaseName']} Template"; wb.properties.description="Controlled blank template; no customer or sample data."
    by_sheet=defaultdict(list)
    for t in mapping["tableMappings"]: by_sheet[t["sheetName"]].append(t)
    raw_sheets={p["sheetName"] for p in mapping["protectedTables"] if p["protection"]=="appendOnly"}
    for sheet_name in mapping["requiredSheetOrder"]:
        ws=wb.create_sheet(sheet_name); heading(ws,mapping["phaseName"],sheet_name,sheet_name in raw_sheets)
        for t in sorted(by_sheet[sheet_name],key=lambda x:(x["headerRow"],x.get("startColumn",1))): add_table(ws,t,mapping,metrics_lookup)
        ws.sheet_properties.tabColor="A5A5A5" if sheet_name in raw_sheets else NAVY if sheet_name.startswith("01_") else BLUE
    output.parent.mkdir(parents=True,exist_ok=True); wb.save(output)


def validate_book(path, expected_sheets):
    wb=load_workbook(path,read_only=False,data_only=False)
    assert wb.sheetnames==expected_sheets,(path,wb.sheetnames)
    names=[]
    for ws in wb.worksheets: names.extend(ws.tables.keys())
    assert len(names)==len(set(names)),(path,"duplicate table names")
    wb.close()


def write_json(path,data):
    path.parent.mkdir(parents=True,exist_ok=True); path.write_text(json.dumps(data,indent=2,ensure_ascii=False)+"\n",encoding="utf-8")


def main():
    specs=phase_specs(); mappings={k:build_map(k,v) for k,v in specs.items()}
    metrics_lookup={
        "tblPreSalesExecutiveEstimate":PS_METRICS,
        "tblPreMigrationExecutiveSummary":PM_METRICS,
        "tblPostMigrationExecutiveSummary":POST_METRICS,
    }
    paths={
        "pre-sales":ROOT/"templates/controlled/pre-sales/eMAS_PreSales_Template.xlsx",
        "pre-migration":ROOT/"templates/controlled/pre-migration/eMAS_PreMigration_Template.xlsx",
        "post-migration":ROOT/"templates/controlled/post-migration/eMAS_PostMigration_Template.xlsx",
    }
    versioned={
        "pre-sales":ROOT/"templates/report-redesign-v3.2/pre-sales/eMAS_PreSales_Template.xlsx",
        "pre-migration":ROOT/"templates/report-redesign-v3.2/pre-migration/eMAS_PreMigration_Template.xlsx",
        "post-migration":ROOT/"templates/report-redesign-v3.2/post-migration/eMAS_PostMigration_Template.xlsx",
    }
    map_paths={
        "pre-sales":ROOT/"config/report-mappings/pre-sales.template-map.json",
        "pre-migration":ROOT/"config/report-mappings/pre-migration.template-map.json",
        "post-migration":ROOT/"config/report-mappings/post-migration.template-map.json",
    }
    for phase,mapping in mappings.items():
        generate_workbook(phase,mapping,metrics_lookup,paths[phase]); validate_book(paths[phase],mapping["requiredSheetOrder"])
        versioned[phase].parent.mkdir(parents=True,exist_ok=True); shutil.copy2(paths[phase],versioned[phase])
        write_json(map_paths[phase],mapping)
        write_json(ROOT/f"config/report-mappings/report-redesign-v3.2/{phase}.template-map.json",mapping)
        write_json(ROOT/f"config/result-schemas/report-redesign-v3.2/{phase}.result.schema.json",result_schema(phase,mapping))
        print(f"Generated {phase}: {paths[phase]}")
    schema=map_schema(); write_json(ROOT/"config/report-mappings/report-template-map.schema.json",schema); write_json(ROOT/"config/report-mappings/report-redesign-v3.2/report-template-map.schema.json",schema)
    readme=ROOT/"templates/report-redesign-v3.2/README.md"; readme.parent.mkdir(parents=True,exist_ok=True); readme.write_text(
        "# eMAS Report Templates — v3.2\n\nGenerated reproducibly by `tools/report-redesign-v3.2/Generate-eMASReportArtifactsV32.py`.\n\n"
        "- Template version: `1.2.0`\n- Report-template map version: `2.0.0`\n- Pre-Sales: 4 sheets\n- Pre-Migration: 11 sheets\n- Post-Migration: 15 sheets\n\n"
        "The workbooks contain controlled blank tables and no customer/sample data. Business interpretation remains in the shared controlled runtime JSON.\n",encoding="utf-8")

if __name__=="__main__": main()
